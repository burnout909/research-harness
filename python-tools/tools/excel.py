"""Excel read/write tools using openpyxl."""
from __future__ import annotations

from pathlib import Path
from typing import Any

import numpy as np
import openpyxl
from scipy.io import loadmat


def excel_read(file_path: str, sheet: str | None = None) -> list[dict[str, Any]]:
    """Read an Excel file and return its contents as a list of row dicts.

    Args:
        file_path: Path to the .xlsx file.
        sheet: Optional sheet name. Defaults to the active sheet.

    Returns:
        List of dicts where keys are column headers from the first row.
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
    return [dict(zip(headers, row)) for row in rows[1:]]


def excel_write(
    file_path: str,
    data: list[dict[str, Any]],
    sheet: str | None = None,
) -> str:
    """Write data to an Excel file.

    Args:
        file_path: Destination .xlsx path.
        data: List of row dicts. Keys become column headers.
        sheet: Optional sheet name.

    Returns:
        The path of the written file.
    """
    if not data:
        raise ValueError("data must be a non-empty list of dicts")

    path = Path(file_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    if sheet:
        ws.title = sheet

    headers = list(data[0].keys())
    ws.append(headers)
    for row in data:
        ws.append([row.get(h) for h in headers])

    wb.save(path)
    wb.close()
    return str(path)


def mat_to_excel(mat_file: str, output_path: str) -> str:
    """Convert a .mat file to .xlsx.

    Each variable in the .mat file becomes a separate sheet.

    Args:
        mat_file: Path to the .mat file.
        output_path: Destination .xlsx path.

    Returns:
        The path of the written Excel file.
    """
    data = loadmat(mat_file, squeeze_me=True)
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    first = True

    for key, val in data.items():
        if key.startswith("_"):
            continue

        ws = wb.active if first else wb.create_sheet(title=key[:31])
        if first:
            ws.title = key[:31]
            first = False

        arr = np.atleast_2d(np.array(val))
        for r_idx, row in enumerate(arr, 1):
            for c_idx, cell in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=float(cell) if np.isscalar(cell) else str(cell))

    if first:
        # No data variables found
        wb.active.title = "empty"

    wb.save(out)
    wb.close()
    return str(out)
