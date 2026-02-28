import openpyxl
import sys

file_path = "../../../data/uploads/1772251815979_2026_________________________2_.xlsx"

try:
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # Iterate backwards from max_row to find the last row with content
    last_row_index = sheet.max_row
    while last_row_index > 0:
        row = sheet[last_row_index]
        has_content = False
        for cell in row:
            if cell.value is not None:
                has_content = True
                break
        if has_content:
            break
        last_row_index -= 1

    # Check if we are at row 0 (empty sheet)
    if last_row_index == 0:
        last_row_index = 1

    print(f"Last row with content is: {last_row_index}")

    # Now append at last_row_index + 1
    target_row = last_row_index + 1

    # Create the cells
    sheet.cell(row=target_row, column=1, value="이동")
    sheet.cell(row=target_row, column=2, value="인천")
    sheet.cell(row=target_row, column=3, value="2024-08-16")

    wb.save(file_path)
    print(f"Successfully appended row at {target_row}")

except Exception as e:
    print(f"Error: {e}")
